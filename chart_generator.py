from logger import logger_decorator
from openpyxl.chart import PieChart, Reference


@logger_decorator
def update_pie_chart(sheet, *args, **kwargs):
    labels = []
    amounts = []

    for row in sheet.iter_rows(min_row=2, max_col=5, max_row=sheet.max_row, values_only=True):
        label = row[2]
        amount = row[4]

        labels.append(label)
        amounts.append(amount)

    label_totals = {}

    for label, amount in zip(labels, amounts):
        if amount is not None:
            if label in label_totals:
                label_totals[label] += amount
            else:
                label_totals[label] = amount

    sheet.cell(row=1, column=10, value="Total Labels")
    sheet.cell(row=1, column=11, value="Amount")
    i = 0

    for label in label_totals:
        total_label = label
        total_amount = label_totals[label]
        sheet.cell(row=2 + i, column=10, value=total_label)
        sheet.cell(row=2 + i, column=11, value=total_amount)
        print("i:" + str(i))
        i += 1

    pie = PieChart()
    pie.title = "Pie Chart"
    # Create a Reference object for the data
    print("max rows = " + str(sheet.max_row))
    labels = Reference(sheet, min_col=10, min_row=2, max_row=sheet.max_row, max_col=10)
    data = Reference(sheet, min_col=11, min_row=1, max_row=sheet.max_row, max_col=11)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    sheet.add_chart(pie, "G1")
    pie.width = 5
    pie.height = 5
    return sheet
