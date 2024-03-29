import tkinter as tk
from tkinter import ttk
from datetime import datetime
from openpyxl import load_workbook
from tkcalendar import DateEntry
from PIL import Image, ImageTk
from chart_generator import update_pie_chart
from logger import logger_decorator


class MoneyManagerApp:
    @logger_decorator
    def __init__(self, *args, **kwargs):

        self.chart_created_for_sheet = {}

        self.root = tk.Tk()
        self.root.title("Money Manger")

        self.load_and_display_image()

        # Create or load workbook
        self.workbook = load_workbook("money_manager.xlsx")

        # GUI components
        self.label_frame = ttk.Frame(self.root)
        self.label_frame.grid(row=0, column=0, padx=10, pady=10)

        self.responsible_options = ["The power couple", "Karina", "Yosef"]
        self.responsible_label = ttk.Label(self.label_frame, text="responsible:")
        self.responsible_label.grid(row=1, column=0, padx=5, pady=5, sticky="w")

        self.responsible_combobox = ttk.Combobox(self.label_frame, values=self.responsible_options, state="readonly")
        self.responsible_combobox.grid(row=1, column=1, padx=5, pady=5)
        self.responsible_combobox.set(self.responsible_options[0])

        self.label_options = ["Car", "Food", "Clothes", "Magics", "Friends & Fun", "Housing", "Finance"]

        self.label_label = ttk.Label(self.label_frame, text="Label:")
        self.label_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")

        self.label_combobox = ttk.Combobox(self.label_frame, values=self.label_options, state="readonly")
        self.label_combobox.grid(row=2, column=1, padx=10, pady=10)
        self.label_combobox.set(self.label_options[0])  # Set default value for combobox

        self.type_label = ttk.Label(self.label_frame, text="Type:")
        self.type_label.grid(row=3, column=0, padx=5, pady=5)

        self.type_entry = ttk.Entry(self.label_frame, width=20)
        self.type_entry.grid(row=3, column=1, padx=5, pady=5)
        self.amount_label = ttk.Label(self.label_frame, text="Amount:")
        self.amount_label.grid(row=4, column=0, padx=5, pady=5)

        self.amount_entry = ttk.Entry(self.label_frame, width=20)
        self.amount_entry.grid(row=4, column=1, padx=5, pady=5)

        self.date_label = ttk.Label(self.label_frame, text="Date:")
        self.date_label.grid(row=5, column=0, padx=5, pady=5)

        self.date_entry = DateEntry(self.label_frame, width=12, background='lightgreen',
                                    foreground='white', borderwidth=2, maxdate=datetime.now())
        self.date_entry.grid(row=5, column=1, padx=10, pady=10)

        self.add_button = ttk.Button(self.label_frame, text="Add", command=self.add_entry)
        self.add_button.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

    @logger_decorator
    def generate_charts(self, active_sheet, *args, **kwargs):

        if active_sheet.max_row <= 2:
            update_pie_chart(active_sheet)
            self.workbook.save("money_manager.xlsx")
        else:
            # Only update the chart without creating a new one
            update_pie_chart(active_sheet)
            self.workbook.save("money_manager.xlsx")

    @logger_decorator
    def load_and_display_image(self, *args, **kwargs):
        background_image = Image.open(
            "C:/Users/josef/PycharmProjects/MoneyManagerApp/pythonProject/IMG_4306.jpg")
        rotated_image = background_image.rotate(-90, expand=True)

        resized_image = rotated_image.resize((800, 600))

        background_photo = ImageTk.PhotoImage(resized_image)

        # Set the initial size of the window
        self.root.geometry(f"{resized_image.width}x{resized_image.height}")

        background_label = tk.Label(self.root, image=background_photo)
        background_label.place(relwidth=1, relheight=1)  # Make the label cover the entire window

        # Attach the background photo to the label to prevent it from being garbage collected
        background_label.image = background_photo

    @logger_decorator
    def add_entry(self, *args, **kwargs):
        label = self.label_combobox.get()
        amount = float(self.amount_entry.get())
        entry_type = self.type_entry.get()
        date = self.date_entry.get_date()
        responsible = self.responsible_combobox.get()
        # Format date as YYYY-MM
        sheet_name = date.strftime("%Y-%m")

        # Check if the sheet exists, if not, create a new one
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(title=sheet_name)
            active_sheet = self.workbook[sheet_name]
            headers = ["Date", "Responsible", "Label", "Type", "Amount"]
            active_sheet.append(headers)
        else:
            active_sheet = self.workbook[sheet_name]

        active_sheet.column_dimensions['B'].width = 18
        active_sheet.column_dimensions['E'].width = 10.14
        active_sheet.column_dimensions['C'].width = 15
        active_sheet.column_dimensions['J'].width = 20
        # Append the entry to the active sheet
        row = [date, responsible, label, entry_type, amount]
        active_sheet.append(row)

        # Save the workbook
        self.workbook.save("money_manager.xlsx")

        # Clear entry fields
        self.amount_entry.delete(0, tk.END)
        self.type_entry.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)

        print("Entry added:", {"Date": date, "Label": label, "Amount": amount, "Type": entry_type})
        print("Current Sheet:")
        for r in active_sheet.iter_rows(min_row=1, max_col=4, max_row=active_sheet.max_row, values_only=True):
            print(r)

        self.generate_charts(active_sheet=active_sheet)

    @logger_decorator
    def run(self, *args, **kwargs):
        self.root.mainloop()


if __name__ == '__main__':
    app = MoneyManagerApp()
    app.run()
